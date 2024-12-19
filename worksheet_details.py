import openpyxl

wb = openpyxl.load_workbook(filename='input.xlsx')

#  CREATE WORKSHEET
wb.create_sheet(index=1, title='Second Sheet')
print(wb.sheetnames)

# DELETE WORKSHEET
del wb['Second Sheet']
print(wb.sheetnames)

# RENAME WORKSHEET
worksheet = wb['Sheet First']
worksheet.title = 'First Sheet'
print(wb.sheetnames)

wb.save(filename='input.xlsx')
