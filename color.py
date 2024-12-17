import openpyxl
from openpyxl.styles import PatternFill

wb = openpyxl.load_workbook(filename='input.xlsx', data_only=True)

sheets = wb.active

yellow = "00FF00"

for i in range(1, sheets.max_row):
    if i % 2 == 0:
        sheets[f'A{i}'].fill = PatternFill(start_color=yellow, end_color=yellow, fill_type='solid')
        sheets[f'B{i}'].fill = PatternFill(start_color=yellow, end_color=yellow, fill_type='solid')
        sheets[f'C{i}'].fill = PatternFill(start_color=yellow, end_color=yellow, fill_type='solid')

wb.save(filename='input.xlsx')
wb.close()
